import openpyxl
from openpyxl.utils.cell import get_column_letter
import datetime
import logging

logging.basicConfig(filename='excel_log.log', level=logging.DEBUG)

file = '/Users/iainrobertson/myCode/Python/ExcelAssignment/expedia_report_monthly_january_2018.xlsx'

def summary_rolling_data(file):
    # print('first data set will be returned as dictionary in here')
    file_year_file_month = get_year_from_filename(file)

    workbook = openpyxl.load_workbook(file)
    worksheet = workbook['Summary Rolling MoM']

    resut_dictionary = {'Calls Offered': 0, 'Abandon after 30s':0, 'FCR':0, 'DSAT':0, "CSAT":0}
    for row in range(1,15):
        for col in range(1,7):
            char = get_column_letter(col) 
            if isinstance(worksheet[char + str(row)].value, datetime.datetime):
                date_time_format = worksheet[char + str(row)].value.strftime("%b") + '-' + worksheet[char + str(row)].value.strftime("%y")

                if date_time_format.lower() == file_year_file_month:
                        resut_dictionary['Calls Offered'] = worksheet[get_column_letter(col+1) + str(row)].value
                        resut_dictionary['Abandon after 30s'] = worksheet[get_column_letter(col+2) + str(row)].value
                        resut_dictionary['FCR'] = worksheet[get_column_letter(col+3) + str(row)].value
                        resut_dictionary['DSAT'] = worksheet[get_column_letter(col+4) + str(row)].value
                        resut_dictionary['CSAT'] = worksheet[get_column_letter(col+5) + str(row)].value
                        break
    return resut_dictionary

def voc_rolling_data(file):
    file_year_file_month = get_year_from_filename(file)

    workbook = openpyxl.load_workbook(file)
    worksheet = workbook['VOC Rolling MoM']
    results = {'Base Size':0, 'Promoters int':0, 'Promoters Percent':0, 'Passives int':0, 'Passives Percent':0, 'Dectractors int':0, 'Dectractors Percent':0, 'Overall NPS':0, 'Sat with Agent':0, 'DSat with Agent':0, 'Promoters good?': '', 'Passives good?':'', 'Dectractors good?': ''}  

    for col in range(1,26):
        char = get_column_letter(col) 

        for row in range(1,23):
            if isinstance(worksheet[char + str(row)].value, datetime.datetime):
                date_time_format = worksheet[char + str(row)].value.strftime("%b") + '-' + worksheet[char + str(row)].value.strftime("%y")
                
                if date_time_format.lower() == file_year_file_month:
                    print(date_time_format.lower())
                    results['Base Size'] = worksheet[char + str(row + 2)].value
                    results['Promoters int'] = worksheet[char + str(row + 3)].value
                    results['Promoters Percent'] = worksheet[char + str(row + 4)].value
                    results['Passives int'] = worksheet[char + str(row + 5)].value
                    results['Passives Percent'] = worksheet[char + str(row + 6)].value
                    results['Dectractors int'] = worksheet[char + str(row + 7)].value
                    results['Dectractors Percent'] = worksheet[char + str(row + 8)].value
                    results['Overall NPS'] = worksheet[char + str(row + 12)].value
                    results['Sat with Agent'] = worksheet[char + str(row + 15)].value
                    results['DSat with Agent'] = worksheet[char + str(row + 18)].value
                    results['Promoters good?'] = 'good' if results['Promoters int'] > 200 else 'bad'
                    results['Passives good?'] = 'good' if results['Passives int'] > 100 else 'bad'
                    results['Dectractors good?'] = 'good' if results['Dectractors int'] > 100 else 'bad'
                    
                    # print(results)
                    break
    return results

def get_year_from_filename(file):
    year = []
    for char in file:
        if char.isdigit():
            year.append(int(char))
    file_year = (year[0] * 1000) + (year[1]*100) + (year[2]*10) + year[3]
    file_month = file[74:77]
    return file_month + '-' + str(file_year)[2:]


try:
    results = summary_rolling_data(file)

    voc_results = voc_rolling_data(file)

    logging.info(datetime.datetime.now())
    logging.info('retrieving summary rolling data from method')
    logging.info(results)
    logging.info('Calls Offered: ' + str(results['Calls Offered']))
    logging.info('Abandon after 30s: ' + str(results['Abandon after 30s'] * 100) + "%")
    logging.info('FCR: ' + str(results['FCR'] * 100) + "%")
    logging.info('DSAT: ' + str(results['DSAT'] * 100) + "%")
    logging.info('CSAT: ' + str(results['CSAT'] * 100) + "%")
    logging.info('-----------------------------------------------')
    logging.info('VOC Dictionary: ' + str(voc_results))
except:
    logging.error('Something has gone wrong')
