import openpyxl
from openpyxl.utils.cell import get_column_letter
import datetime
import logging

logging.basicConfig(filename='excel_log.log', level=logging.DEBUG)

file = '/Users/iainrobertson/myCode/Python/ExcelAssignment/expedia_report_monthly_january_2018.xlsx'
year = []
for char in file:
    if char.isdigit():
        year.append(int(char))
# print(year)
file_year = (year[0] * 1000) + (year[1]*100) + (year[2]*10) + year[3]
# print(file_year)
file_month = file[74:77]
# print(file_month)

file_year_file_month = file_month + '-' + str(file_year)[2:]
print(file_year_file_month)



workbook = openpyxl.load_workbook(file)
worksheet = workbook['Summary Rolling MoM']
worksheet_2 = workbook['VOC Rolling MoM']
cell_value = worksheet['B12'].value

# print(workbook.sheetnames)
# print(cell_value) 

for row in range(1,15):
    for col in range(1,7):
        char = get_column_letter(col) 
        # print(worksheet[char + str(row)].value)   


resut_dictionary = {'Calls Offered': 0, 'Abandon after 30s':0, 'FCR':0, 'DSAT':0, "CSAT":0}
# data I want
for row in range(1,15):
    for col in range(1,7):
     
        char = get_column_letter(col) 

        if isinstance(worksheet[char + str(row)].value, datetime.datetime):
            date_time_format = worksheet[char + str(row)].value.strftime("%b") + '-' + worksheet[char + str(row)].value.strftime("%y")
        

            if date_time_format.lower() == file_year_file_month:
                     
                        resut_dictionary['Calls Offered'] = worksheet[get_column_letter(col+1) + str(row)].value
                        resut_dictionary['Abandon after 30s'] = worksheet[get_column_letter(col+2) + str(row)].value
                        resut_dictionary['FCR'] = worksheet[get_column_letter(col+3) + str(row)].value
                        resut_dictionary['DSAT'] = worksheet[get_column_letter(col+4) + str(row)].value
                        resut_dictionary['CSAT'] = worksheet[get_column_letter(col+5) + str(row)].value

                        logging.info(datetime.datetime.now())
                        logging.info(resut_dictionary)
                        logging.info('Calls Offered: ' + str(resut_dictionary['Calls Offered']))
                        logging.info('Abandon after 30s: ' + str(resut_dictionary['Abandon after 30s'] * 100) + "%")
                        logging.info('FCR: ' + str(resut_dictionary['FCR'] * 100) + "%")
                        logging.info('DSAT: ' + str(resut_dictionary['DSAT'] * 100) + "%")
                        logging.info('CSAT: ' + str(resut_dictionary['CSAT'] * 100) + "%")
                        
                        

                        break
